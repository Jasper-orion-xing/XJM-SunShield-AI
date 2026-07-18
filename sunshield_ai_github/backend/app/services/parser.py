from __future__ import annotations

import io
import re
import zipfile
from dataclasses import dataclass
from html import unescape
from pathlib import Path
from xml.etree import ElementTree


SUPPORTED_EXTENSIONS = {".txt", ".md", ".csv", ".pdf", ".xlsx", ".docx"}


@dataclass(frozen=True)
class ParsedDocument:
    filename: str
    file_type: str
    text: str
    warnings: list[str]
    content: bytes | None = None


def parse_upload(filename: str, content: bytes) -> ParsedDocument:
    suffix = Path(filename).suffix.lower()
    if suffix not in SUPPORTED_EXTENSIONS:
        supported = ", ".join(sorted(SUPPORTED_EXTENSIONS))
        raise ValueError(f"暂不支持 {suffix or '无扩展名'} 文件。支持格式：{supported}")
    if suffix in {".txt", ".md", ".csv"}:
        return ParsedDocument(filename, suffix.lstrip("."), _decode_text(content), [], content)
    if suffix == ".pdf":
        return ParsedDocument(filename, "pdf", _extract_pdf(content), [], content)
    if suffix == ".xlsx":
        return ParsedDocument(filename, "xlsx", _extract_xlsx(content), [], content)
    if suffix == ".docx":
        text, warnings = _extract_docx(content)
        return ParsedDocument(filename, "docx", text, warnings, content)
    raise ValueError("不支持的文件格式")


def _decode_text(content: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ValueError("无法识别文本编码，请转换为 UTF-8 后重试。")


def _extract_pdf(content: bytes) -> str:
    try:
        import fitz
    except ImportError as exc:
        raise RuntimeError("当前环境缺少 PyMuPDF，无法解析 PDF。") from exc

    pages: list[str] = []
    with fitz.open(stream=content, filetype="pdf") as doc:
        for index, page in enumerate(doc, start=1):
            page_text = page.get_text("text").strip()
            if page_text:
                pages.append(f"[PDF 第 {index} 页]\n{page_text}")
    return "\n\n".join(pages)


def _extract_xlsx(content: bytes) -> str:
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("当前环境缺少 openpyxl，无法解析 XLSX。") from exc

    workbook = openpyxl.load_workbook(io.BytesIO(content), data_only=False, read_only=True)
    parts: list[str] = []
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        parts.append(f"[工作表: {sheet_name}]")
        for row in sheet.iter_rows():
            values = []
            for cell in row:
                if cell.value is None:
                    continue
                values.append(f"{cell.coordinate}={cell.value}")
            if values:
                parts.append(" | ".join(values))
    workbook.close()
    return "\n".join(parts)


def _extract_docx(content: bytes) -> tuple[str, list[str]]:
    warnings = ["DOCX 当前解析正文、表格、页眉页脚文本；下载先提供脱敏 TXT。"]
    with zipfile.ZipFile(io.BytesIO(content)) as archive:
        xml_members = [
            name
            for name in archive.namelist()
            if name.startswith("word/")
            and name.endswith(".xml")
            and (
                name == "word/document.xml"
                or name.startswith("word/header")
                or name.startswith("word/footer")
            )
        ]
        chunks: list[str] = []
        for member in sorted(xml_members):
            xml_text = archive.read(member).decode("utf-8", errors="ignore")
            chunks.extend(_docx_text_from_xml(xml_text))
    return "\n".join(chunk for chunk in chunks if chunk.strip()), warnings


def _docx_text_from_xml(xml_text: str) -> list[str]:
    try:
        root = ElementTree.fromstring(xml_text)
    except ElementTree.ParseError:
        fallback = re.sub(r"<[^>]+>", " ", xml_text)
        return [unescape(re.sub(r"\s+", " ", fallback)).strip()]

    namespace = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    paragraphs: list[str] = []
    for paragraph in root.findall(".//w:p", namespace):
        texts = [node.text or "" for node in paragraph.findall(".//w:t", namespace)]
        if texts:
            paragraphs.append("".join(texts))
    return paragraphs
